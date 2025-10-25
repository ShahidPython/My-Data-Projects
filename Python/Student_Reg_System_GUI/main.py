from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.resizable(False, False)
root.config(bg=background)

# Initialize gender variable
gender = ""

# Create directories if they don't exist
os.makedirs(os.path.join(BASE_DIR, "assets", "Student Images"), exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "assets", "Images"), exist_ok=True)

file_path = os.path.join(BASE_DIR, "Student_data.xlsx")

# Initialize the Excel file properly
if not os.path.exists(file_path):
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "D.O.B"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    wb.save(file_path)

####################### Exit window #####################
def Exit():
    root.destroy()

####################### Show Image ####################### 
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(), 
        title="Select image file", 
        filetypes=(
            ("JPG File", "*.jpg"),
            ("PNG File", "*.png"), 
            ("All Files", "*.*")
        )
    )
    
    if filename:
        try:
            img = Image.open(filename)
            resized_image = img.resize((190,190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load image: {str(e)}")

####################### Registration NO. ####################### 
def registration_no():
    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    row = sheet.max_row
    
    # Handle case where sheet only has headers
    if row == 1:
        Registration.set(1)
    else:
        max_row_value = sheet.cell(row=row, column=1).value
        try:
            Registration.set(max_row_value + 1)
        except:
            Registration.set(1)

################ Clear #######################
def Clear():
    global img
    Name.set("")
    DOB.set("")
    Religion.set("")
    Skill.set("")
    F_Name.set("")
    M_Name.set("")
    Father_Occupation.set("")
    Mother_Occupation.set("")
    Class.set("Select Class")
    registration_no()
    saveButton.config(state="normal")
    img1 = PhotoImage(file=os.path.join(BASE_DIR, "assets", "Images", "upload photo.png"))
    lbl.config(image=img1)
    lbl.image = img1
    img = None  # Reset the image variable
    radio.set(0)  # Clear gender selection

###################### Save #######################
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    
    # Check if gender is selected
    if not radio.get():
        messagebox.showerror("error", "Please select Gender!")
        return

    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or Rel == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("error", "Please fill all the fields!")
        return

    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
    sheet.cell(column=2, row=sheet.max_row, value=N1)
    sheet.cell(column=3, row=sheet.max_row, value=C1)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=D2)
    sheet.cell(column=6, row=sheet.max_row, value=D1)
    sheet.cell(column=7, row=sheet.max_row, value=Rel)
    sheet.cell(column=8, row=sheet.max_row, value=S1)
    sheet.cell(column=9, row=sheet.max_row, value=fathername)
    sheet.cell(column=10, row=sheet.max_row, value=mothername)
    sheet.cell(column=11, row=sheet.max_row, value=F1)
    sheet.cell(column=12, row=sheet.max_row, value=M1)
    
    try:
        file.save(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {str(e)}")
        return

    if 'img' in globals() and img is not None:
        try:
            img_path = os.path.join(BASE_DIR, "assets", "Student Images", f"{R1}.jpg")
            img.save(img_path)
        except Exception as e:
            messagebox.showwarning("Warning", f"Profile picture not saved: {str(e)}")
    else:
        messagebox.showwarning("Warning", "No profile picture was uploaded")

    messagebox.showinfo("info", "Data Saved Successfully")
    Clear()
    registration_no()

##################### Search ######################
def search():
    text = Search.get().strip()
    
    if not text:
        messagebox.showerror("Error", "Please enter a registration number to search")
        return

    try:
        search_num = int(text)
    except ValueError:
        messagebox.showerror("Error", "Registration number must be a number")
        return

    Clear()
    saveButton.config(state="disabled")

    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    found = False

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == search_num:
            found = True
            Registration.set(row[0].value)
            Name.set(row[1].value)
            Class.set(row[2].value)
            
            if row[3].value == "Female":
                R2.select()
            else:
                R1.select()
                
            DOB.set(row[4].value)
            Date.set(row[5].value)
            Religion.set(row[6].value)
            Skill.set(row[7].value)
            F_Name.set(row[8].value)
            M_Name.set(row[9].value)
            Father_Occupation.set(row[10].value)
            Mother_Occupation.set(row[11].value)

            # Load student image if exists
            img_path = os.path.join(BASE_DIR, "assets", "Student Images", f"{search_num}.jpg")
            if os.path.exists(img_path):
                try:
                    img = Image.open(img_path)
                    resized_image = img.resize((190,190))
                    photo2 = ImageTk.PhotoImage(resized_image)
                    lbl.config(image=photo2)
                    lbl.image = photo2
                except Exception as e:
                    messagebox.showwarning("Warning", f"Couldn't load student image: {str(e)}")
            break

    if not found:
        messagebox.showerror("Invalid", "Invalid registration number!!!")

################## Update ######################
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    
    # Check if gender is selected
    if not radio.get():
        messagebox.showerror("error", "Please select Gender!")
        return

    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or Rel == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("error", "Please fill all the fields!")
        return

    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    found = False

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == R1:
            row[1].value = N1
            row[2].value = C1
            row[3].value = gender
            row[4].value = D2
            row[5].value = D1
            row[6].value = Rel
            row[7].value = S1
            row[8].value = fathername
            row[9].value = mothername
            row[10].value = F1
            row[11].value = M1
            found = True
            break

    if not found:
        messagebox.showerror("Error", "Registration number not found!")
        return

    try:
        file.save(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to update data: {str(e)}")
        return

    if 'img' in globals() and img is not None:
        try:
            img_path = os.path.join(BASE_DIR, "assets", "Student Images", f"{R1}.jpg")
            img.save(img_path)
        except Exception as e:
            messagebox.showwarning("Warning", f"Profile picture not updated: {str(e)}")

    messagebox.showinfo("Update", "Data Updated Successfully")
    Clear()

################## Gender Selection ######################
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

# top frames
Label(root,text="Email:python1234@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg="#fff",font="arial 20 bold").pack(side=TOP,fill=X)

# search box to update
Search = StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)

imageicon3 = PhotoImage(file=os.path.join(BASE_DIR, "assets", "Images", "search.png"))
Srch = Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg='#68ddfa',font="arial 13 bold",command=search)
Srch.place(x=1060,y=64)

imageicon4 = PhotoImage(file=os.path.join(BASE_DIR, "assets", "Images", "Layer 4.png"))
Update_button = Button(root,image=imageicon4,bg="#c36464",command=Update)
Update_button.place(x=110,y=64)

# Registration and Date
Label(root,text="Registration No:",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)

# Student details
obj = LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name = StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB = StringVar()
dob_entry = Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

radio = IntVar()
R1 = Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2 = Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=250,y=150)

Religion = StringVar()
religion_entry = Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion_entry.place(x=630,y=100)

Skill = StringVar()
skill_entry = Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=630,y=150)

Class = Combobox(obj,values=["1","2","3","4","5","6","7","8","9","10","11","12"],font="Roboto 10",width=17,state='r')
Class.place(x=630,y=50)
Class.set("Select Class")

# Parents details
obj2 = LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

Label(obj2,text="Mother's Name",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

F_Name = StringVar()
f_entry = Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)

M_Name = StringVar()
M_entry = Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
M_entry.place(x=630,y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)

# image
f = Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

# Default image
default_image_path = os.path.join(BASE_DIR, "assets", "Images", "upload photo.png")
if not os.path.exists(default_image_path):
    # Create a blank image if default doesn't exist
    img = Image.new('RGB', (190, 190), color='black')
    img.save(default_image_path)

img = PhotoImage(file=default_image_path)
lbl = Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

# button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)

root.mainloop()