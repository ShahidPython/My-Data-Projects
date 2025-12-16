from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

root = Tk()
root.title("Data Entry")
root.geometry("700x400+300+200")
root.resizable(False, False)
root.configure(bg="#326273")

file_path = os.path.join(BASE_DIR, 'Backened_data.xlsx')
file = pathlib.Path(file_path)
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Contact Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"

    file.save(file_path)

def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save(file_path)

    messagebox.showinfo("info", "detail added!")
    nameValue.set("")
    contactValue.set("")
    AgeValue.set("")
    addressEntry.delete(1.0, END)
    
def clear():
    nameValue.set("")
    contactValue.set("")
    AgeValue.set("")
    addressEntry.delete(1.0, END)

# icon
icon_image = PhotoImage(file=os.path.join(BASE_DIR, "assets/logo.png"))
root.iconphoto(False, icon_image)

# heading
Label(root, text="Please fill out this Entry Form:", font="arial 13",bg = "#326273",fg = "#fff").place(x=20,y=20)

# label
Label(root, text="Name", font=23,bg = "#326273",fg = "#fff").place(x=50,y=100)
Label(root, text="Contact No.", font=23,bg = "#326273",fg = "#fff").place(x=50,y=150)
Label(root, text="Age",font=23,bg = "#326273",fg = "#fff").place(x=50,y=200)
Label(root, text="Gender", font=23,bg = "#326273",fg = "#fff").place(x=370,y=200)
Label(root, text="Address", font=23,bg = "#326273",fg = "#fff").place(x=50,y=250)

# Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width = 45,bd = 2,font = 20)
contactEntry = Entry(root, textvariable=contactValue, width = 45,bd = 2,font = 20)
ageEntry = Entry(root, textvariable=AgeValue, width = 15,bd = 2,font = 20)

# gender
gender_combobox = Combobox(root, values=["Male", "Female", "Other"], width = 14,state='r', font = 'arial 14')
gender_combobox.place(x=440,y=200)
gender_combobox.set("Male")

addressEntry = Text(root, width = 50, height = 4,bd = 2)

nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

# Button Style Configuration
button_style = {
    "font": ("Arial", 12, "bold"),
    "width": 12,
    "height": 1,
    "bd": 2,
    "relief": "raised",
    "cursor": "hand2"
}

# Fantastic Buttons moved down (y=340 instead of y=320)
Button(root, text="Submit", bg="#4CAF50", fg="white", 
       activebackground="#45a049", activeforeground="white",
       **button_style, command=submit).place(x=200,y=340)

Button(root, text="Clear", bg="#FF9800", fg="white", 
       activebackground="#e68900", activeforeground="white",
       **button_style, command=clear).place(x=340,y=340)

Button(root, text="Exit", bg="#f44336", fg="white", 
       activebackground="#da190b", activeforeground="white",
       **button_style, command=lambda:root.destroy()).place(x=480,y=340)

root.mainloop()